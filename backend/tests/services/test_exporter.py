from __future__ import annotations

import json
from pathlib import Path

import pytest

from app.services.exporter import (
    gather_approved_rows,
    export_csv,
    export_xlsx,
    DEFAULT_HEADERS,
)


def write_audit(tmp: Path, bank: str, file_id: str, payload: dict) -> Path:
    bank_dir = tmp / bank
    bank_dir.mkdir(parents=True, exist_ok=True)
    p = bank_dir / f"{file_id}.json"
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return p


def base_payload(bank: str, file_id: str) -> dict:
    return {
        "schema_version": 1,
        "generated_at": "2025-01-01T00:00:00Z",
        "bank": bank,
        "file": file_id,
        "decision": {
            "decision": "review",
            "stp": False,
            "overall_conf": 0.95,
            "low_conf_fields": [],
            "reasons": [],
        },
        "fields": {
            "date": {
                "parse_norm": "2025-01-31",
                "ocr_text": "31/Jan/2025",
                "validation": {"ok": True},
            },
            "cheque_number": {"parse_norm": "12345678", "ocr_text": "12345678", "validation": {"ok": True}},
            "amount_numeric": {"parse_norm": "100.00", "ocr_text": "100.00", "validation": {"ok": True}},
            "name": {"parse_norm": "شركة بالم هيلز للتعمير", "ocr_text": "شركة بالم هيلز للتعمير", "validation": {"ok": True}},
        },
    }


def test_gather_approved_rows_filters_and_builds(tmp_path: Path):
    # Not approved (review, stp False) -> excluded
    p1 = base_payload("FABMISR", "f1.jpg")
    write_audit(tmp_path, "FABMISR", "f1.jpg", p1)

    # Approved by stp True -> included
    p2 = base_payload("FABMISR", "f2.jpg")
    p2["decision"]["stp"] = True
    write_audit(tmp_path, "FABMISR", "f2.jpg", p2)

    # Approved by decision auto_approve -> included
    p3 = base_payload("QNB", "q1.jpg")
    p3["decision"]["decision"] = "auto_approve"
    write_audit(tmp_path, "QNB", "q1.jpg", p3)

    # Missing validation.ok True on a required field -> excluded
    p4 = base_payload("QNB", "q2.jpg")
    p4["fields"]["name"]["validation"] = {"ok": False}
    write_audit(tmp_path, "QNB", "q2.jpg", p4)

    rows = gather_approved_rows(str(tmp_path))
    keys = {(r.bank, r.file) for r in rows}
    assert keys == {("FABMISR", "f2.jpg"), ("QNB", "q1.jpg")}

    # Spot-check row data
    row_map = { (r.bank, r.file): r for r in rows }
    r = row_map[("QNB", "q1.jpg")]
    assert r.date == "2025-01-31"
    assert r.cheque_number == "12345678"
    assert r.amount_numeric == "100.00"
    assert r.name.startswith("شركة")


def test_export_csv_and_xlsx(tmp_path: Path):
    rows = gather_approved_rows(str(tmp_path))
    assert rows == []  # empty when no data

    # Create one approved row
    p = base_payload("FABMISR", "f3.jpg")
    p["decision"]["stp"] = True
    write_audit(tmp_path, "FABMISR", "f3.jpg", p)

    rows = gather_approved_rows(str(tmp_path))
    assert len(rows) == 1

    csv_path = tmp_path / "out" / "cheques.csv"
    out_csv = export_csv(str(csv_path), rows)
    content = Path(out_csv).read_text(encoding="utf-8").splitlines()
    assert content[0].split(",") == list(DEFAULT_HEADERS)
    assert content[1].split(",")[0:2] == ["FABMISR", "f3.jpg"]

    # XLSX (skip if openpyxl not installed in test env)
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        pytest.skip("openpyxl not installed")

    xlsx_path = tmp_path / "out" / "cheques.xlsx"
    out_xlsx = export_xlsx(str(xlsx_path), rows)
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(out_xlsx)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == list(DEFAULT_HEADERS)
    data_row = [c.value for c in ws[2]][0:2]
    assert data_row == ["FABMISR", "f3.jpg"]
